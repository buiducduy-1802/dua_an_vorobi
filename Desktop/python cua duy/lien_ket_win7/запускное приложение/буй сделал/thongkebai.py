import os
import sys
import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from functools import partial
from tkinter import ttk
from tkinter import filedialog
from PIL import Image, ImageTk

bg_color = 'darkgreen'
button_color = "greenlight"

# Lấy thư mục hiện tại, linh hoạt cho cả file .py và .exe
if getattr(sys, 'frozen', False):  # Kiểm tra nếu đang chạy từ file .exe
    current_folder = os.path.dirname(sys.executable)  # Thư mục của file .exe
else:
    current_folder = os.path.dirname(os.path.abspath(__file__))  # Thư mục của file .py

# Trỏ đến thư mục chứa file Excel
excel_folder = current_folder
def create_excel_buttons():
    
    global file_path
    for widget in file_button_frame.winfo_children():
        widget.destroy()
    for file in os.listdir(excel_folder):
        if file.endswith(".xlsx"):
            file_path = os.path.join(excel_folder, file)
            file_name = file.rsplit(".", 1)[0]  # Lấy tên tệp mà không có phần mở rộng
            btn = tk.Button(file_button_frame, text=file_name, command=partial(select_file, file_path),width=12,wraplength=100)
            btn.grid(sticky="w", padx=5, pady=5)
    
    
    file_button_frame.update_idletasks()
    canvas1.config(scrollregion=canvas1.bbox("all"))

def select_file(file_path):
    global selected_file, text_id, workbook, image_label, entries
    entries = []  # Đảm bảo biến được khởi tạo
    selected_file = file_path
    
    if file_path:
        file_name = os.path.splitext(os.path.basename(file_path))[0]
        # Xóa văn bản cũ nếu đã tồn tại
        if 'text_id' in globals() and text_id is not None:
            fixed_canvas.delete(text_id)
        text_id = fixed_canvas.create_text(10, 10, text=file_name, anchor="nw", font=("Arial", 12, "bold"), fill='white')
        
    # Xóa widget cũ
    for widget in row_button_frame.winfo_children():
        widget.destroy()
    for widget in row_data_frame.winfo_children():
        widget.destroy()
    for widget in edit_button_frame.winfo_children():
        widget.destroy()

    try:
        global sheet
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # Thêm nhãn và hiển thị ảnh
        image_label = tk.Label(row_data_frame, text="Ảnh sẽ hiển thị tại đây", bg="lightgrey", width=300, height=300)
        image_label.place(x=10, y=30)
        load_image_from_excel()

        # Nhãn lớn
        label = tk.Label(
            row_data_frame,
            text="характеристика оружия",
            font=("Arial", 16, "bold"),
            fg="seagreen",
            bg="white",
            padx=10,
            pady=10
        )
        label.place(x=350, y=10)

        # Tạo Text box
        text_box_canvas = tk.Frame(row_data_frame)
        text_box_canvas.place(x=350, y=50)
        text_box = tk.Text(text_box_canvas, width=60, height=28, wrap="word")
        text_box.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = tk.Scrollbar(text_box_canvas, command=text_box.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        text_box.config(yscrollcommand=scrollbar.set)

        # Lấy dữ liệu từ Excel
        value = sheet["H8"].value if sheet["H8"].value is not None else"" 
        text_box.insert("1.0", str(value))
        entries.append(text_box)
        
        def edit_row_data_fixed():

            try:
                # Kiểm tra xem file Excel đã được chọn hay chưa
                if not selected_file:
                    messagebox.showerror("Ошибка", "Выберите файл Excel перед редактированием!")
                    return

                # Tải file Excel
                workbook = load_workbook(selected_file)
                sheet = workbook.active

                # Lấy dữ liệu từ Text Box
                if isinstance(text_box, tk.Text):
                    value = text_box.get("1.0", "end-1c").strip()  # Lấy nội dung từ Text Box
                else:
                    messagebox.showerror("Ошибка", "Текстовое поле не найдено!")
                    return

                # Cập nhật ô H8
                sheet["H8"] = value

                # Lưu và đóng file Excel
                workbook.save(selected_file)
                workbook.close()

                # Thông báo thành công
                messagebox.showinfo("успех", "Ячейка H8 успешно обновлена!")
            except Exception as e:
                # Hiển thị thông báo lỗi nếu có
                messagebox.showerror("ошибка", f"Не удалось обновить ячейку H8: {e}")

        nut_cn=tk.Button(row_data_frame,text="обновить данны",command=edit_row_data_fixed)
        nut_cn.place(x=180,y=350)
        # Tạo nút cho từng hàng
        for row_num in range(1, sheet.max_row + 1):
            if any(cell.value is not None for cell in sheet[row_num]):
                button_text = sheet.cell(row=row_num, column=1).value or f"Row {row_num}"
                btn = tk.Button(
                    row_button_frame,
                    text=button_text,
                    command=lambda rn=row_num: show_row_data(rn),
                    width=14,
                    wraplength=100
                )
                btn.grid(row=row_num - 1, column=0, sticky="w", padx=5, pady=5)
        
        # Các nút chỉnh sửa
        add_data_btn = tk.Button(edit_button_frame, text="добовить упражнения", command=prepare_new_row)
        add_data_btn.grid(row=0, column=0, padx=5, pady=5)

        sua_data_btn = tk.Button(edit_button_frame, text="исправить упражнения", command=edit_row_data)
        sua_data_btn.grid(row=0, column=1, padx=5, pady=5)

        clear_row_btn = tk.Button(edit_button_frame, text="удаление упражнения", command=clear_row)
        clear_row_btn.grid(row=0, column=2, padx=5, pady=5)
        
                # Nút cập nhật ảnh
        btn_update_image = tk.Button(edit_button_frame, text="менять фото", command=update_image)
        btn_update_image.grid(row=0, column=3, padx=5, pady=5)

    except Exception as e:
        messagebox.showerror("ошибка", f"не нашел оружие: {e}")


def load_image_from_excel():
    """Hàm để tải và hiển thị ảnh từ Excel."""
    try:
        image_path = sheet.cell(row=2, column=7).value  # Đường dẫn ảnh trong ô B2
        if image_path and os.path.exists(image_path):
            img = Image.open(image_path)
            img = img.resize((300, 300))  # Resize ảnh
            img_tk = ImageTk.PhotoImage(img)
            image_label.config(image=img_tk)
            image_label.image = img_tk
        else:
            messagebox.showwarning("Предупреждение", "Изображение не найдено по пути")
            image_label.config(image='', text="Фотография не существует или не выбрана.")
    except Exception as e:
        messagebox.showerror("Ошибка", f"Невозможно загрузить фото: {e}")


def update_image():
    """Hàm để thay đổi và cập nhật đường dẫn ảnh trong Excel."""
    new_image_path = filedialog.askopenfilename(
        title="Выбрать новую фотографию",
        filetypes=[("Image files", "*.jpg *.png *.jpeg *.bmp *.gif")]
    )
    if new_image_path:
        try:
            base_dir = os.path.dirname(selected_file)  # Lấy thư mục của file Excel
            relative_image_path = os.path.relpath(new_image_path, base_dir) 
            # Lưu đường dẫn ảnh mới vào Excel
            sheet.cell(row=2, column=7).value = relative_image_path
            workbook.save(selected_file)  # Lưu lại vào file Excel gốc
            messagebox.showinfo("Уведомление", "Фотография успешно обновлена!")
            load_image_from_excel()  # Tải ảnh mới lên giao diện
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể cập nhật ảnh: {e}")
            
def show_row_data(row_num):
    global selected_row, entries
    selected_row = row_num

    try:
        workbook = load_workbook(selected_file)
        sheet = workbook.active
        row_data = [cell.value for cell in sheet[row_num]]

        # Xóa các widget hiện có trong `entries` và khung `row_data_frame`
        for widget in row_data_frame.winfo_children():
            widget.destroy()
        entries.clear()  # Làm rỗng danh sách entries trước khi tạo lại

        # Tạo các label và entry cho từng ô trong hàng đã chọn
        labels = ["имя упражнения", "времия создания", "количество курсанты", "содержание управнение"]
        for col in range(4):
            if col < 1:
                entry = tk.Entry(row_data_frame, width=50)
                entry.grid(row=1, column=col, padx=5, pady=5)
                entry.insert(0, row_data[col] if row_data[col] is not None else "")
                entries.append(entry)
                label = tk.Label(row_data_frame, text=labels[col])
                label.grid(row=0, column=col, padx=5, pady=5)
            
            elif col < 3:  # Các Entry cho cột đầu tiên đến cột thứ ba
                entry = tk.Entry(row_data_frame, width=20)
                entry.grid(row=1, column=col, padx=5, pady=5)
                entry.insert(0, row_data[col] if row_data[col] is not None else "")
                entries.append(entry)
                label = tk.Label(row_data_frame, text=labels[col])
                label.grid(row=0, column=col, padx=5, pady=5)
            else :  # Text box cho cột thứ tư
                label = tk.Label(row_data_frame, text=labels[col])
                label.grid(row=2, column=col-3, padx=5, pady=5)
                text_box_cavas = tk.Text(row_data_frame, width=75, height=25)
                text_box_cavas.grid(row=3, column=col-3,columnspan=2, padx=10, pady=10)
                text_box = tk.Text(text_box_cavas, width=65, height=25, wrap="word")
                text_box.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
                text_box.insert("1.0", row_data[col] if row_data[col] is not None else "")
                entries.append(text_box)
                # Tạo Scrollbar
                scrollbar = tk.Scrollbar(text_box_cavas, command=text_box.yview)  # Điều khiển theo chiều dọc
                scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
                # Liên kết Text với Scrollbar
                text_box.config(yscrollcommand=scrollbar.set)
            
            dan_duoc=tk.Frame(row_data_frame,highlightbackground=bg_color, highlightthickness=3)
            dan_duoc.place(x=570, y=80, width=300, height=440)
            loai_dan=tk.Label(dan_duoc,text="вид боеприпасов")
            loai_dan.grid(row=1, column=1, sticky="n", padx=5, pady=5)
            so_dan_moi_nguoi=tk.Label(dan_duoc,text="количество боеприпасов")
            so_dan_moi_nguoi.grid(row=1, column=2, sticky="e", padx=5, pady=5)
            entry_list_1 = []
            def update_sum():
                total = 0
                for entry in entry_list_1:
                    try:
                        ammo_per_person = int(row_data[2])
                        total = total+ float(entry.get())*ammo_per_person if entry.get() else 0
                    except ValueError:
                        pass
                total_label.config(text=f"общее количество: {total}")
            row_indices = [row_num_d for row_num_d in range(1, sheet.max_row + 1) if sheet.cell(row=row_num_d, column=10).value is not None]

            for index, row_num_d in enumerate(row_indices, start=1):
                button_text_1 = sheet.cell(row=row_num_d, column=10).value
                btn_1 = tk.Button(
                    dan_duoc,
                    text=button_text_1,
                    wraplength=100,
                    width=20
                )
                btn_1.grid(row=index+1, column=1, sticky="e", padx=5, pady=5)
                
                    # Entry để nhập số
                entry_1 = tk.Entry(dan_duoc, width=10)
                entry_1.grid(row=index+1, column=2, padx=5, pady=5)
                entry_1.bind("<KeyRelease>", lambda _: update_sum())
                
                # Lưu Entry vào danh sách để tính tổng
                entry_list_1.append(entry_1)
            total_label = tk.Label(dan_duoc, text="общее количество: 0", font=("Arial", 12, "bold"))
            total_label.grid(row=len(row_indices) + 2, column=1, columnspan=2, pady=10)
            them_dan=tk.Frame(root,bg='darkgreen')
            them_dan.pack(pady=5)
            def show_entry():
                # Khi nhấn nút "Thêm dữ liệu", ô Entry sẽ xuất hiện
                them_dan.pack(pady=5)
                label_entry.pack(pady=5)
                entry.pack(pady=5)
                confirm_button.pack(pady=5)

                # Ẩn nút "Thêm dữ liệu"
                add_button.pack_forget()
            def add_data_to_excel():
                # Lấy dữ liệu từ entry widget
                data = entry.get()
                try:
                    workbook = load_workbook(selected_file)
                    sheet = workbook.active
                except FileNotFoundError:
                    workbook = load_workbook()
                    sheet = workbook.active

                # Tìm hàng đầu tiên trống trong cột 10 (Cột J)
                for row in range(1, sheet.max_row + 1):
                    if sheet.cell(row=row, column=10).value is None:
                        sheet.cell(row=row, column=10).value = data
                        break
                else:
                    # Nếu không tìm thấy hàng trống, thêm vào hàng mới
                    sheet.append([None] * 9 + [data])

                # Lưu lại file Excel
                workbook.save(selected_file)
                show_row_data(row_num)
                    # Xóa dữ liệu trong entry widget và ẩn ô entry
                entry.delete(0, tk.END)
                label_entry.pack_forget()
                entry.pack_forget()
                confirm_button.pack_forget()
                them_dan.pack_forget()

                # Hiển thị thông báo thành công
                messagebox.showinfo("Успех", "Данные добавлены!")

                # Hiện lại nút "Thêm dữ liệu"
                add_button.pack(pady=10)
            # Tạo nút "Thêm dữ liệu"
            add_button = tk.Button(row_data_frame, text="Больше данных", command=show_entry)
            add_button.place(x=760,y=30)

            # Tạo widget nhập dữ liệu (sẽ ẩn ban đầu)

            label_entry = tk.Label(them_dan, text="Введите вид боеприпасов:")
            entry = tk.Entry(them_dan)
            confirm_button = tk.Button(them_dan, text="Дальнейшее подтверждение", command=add_data_to_excel)
            them_dan.pack_forget()
                    
                    
        # Tạo LabelFrame cho phần bổ sung
        # calc_frame = tk.LabelFrame(row_data_frame, text="Посчитайте количество пуль", padx=10, pady=10)
        # calc_frame.grid(row=3, column=2, columnspan=5, padx=10, pady=10, sticky="w")

        # # Thêm Entry cho Количество участников và Label cho kết quả
        # tk.Label(calc_frame, text="Количество участников").grid(row=0, column=0, padx=5, pady=5)
        # num_people_entry = tk.Entry(calc_frame, width=10)
        # num_people_entry.grid(row=1, column=0, padx=5, pady=5)

        # tk.Label(calc_frame, text="Необходимое количество пуль").grid(row=2, column=0, padx=5, pady=5)
        # result_label = tk.Label(calc_frame, text="", width=10, bg="lightgrey")
        # result_label.grid(row=2, column=1, padx=5, pady=5)

        # # Hàm Посчитайте количество пуль cần dùng
        # def calculate_ammo():
        #     try:
        #         num_people = int(num_people_entry.get())
        #         ammo_per_person = int(row_data[2])  # Lấy giá trị "количество пуля каждого человека" từ dữ liệu Excel
        #         total_ammo = num_people * ammo_per_person
        #         result_label.config(text=str(total_ammo))
        #     except ValueError:
        #         messagebox.showerror("Ошибка", "Пожалуйста, введите действительный номер")

        # # Nút để расчет
        # calc_button = tk.Button(calc_frame, text="читать", command=calculate_ammo)
        # calc_button.grid(row=1, column=1, padx=5, pady=5)

    except Exception as e:
        messagebox.showerror("ошибка", f"не нашел данны {e}")


def prepare_new_row():
    global selected_row, entries
    selected_row = None  # Đặt lại selected_row để biểu thị hàng mới
    entries = []  # Xóa entries cũ để lưu các widget mới

    # Xóa dữ liệu trong các Text widget để tạo hàng mới
    for widget in row_data_frame.winfo_children():
        widget.destroy()

    labels = ["имя упражнения", "времия создания", "количество пуля каждого человека", "содержание управнение"]
    
    for col in range(4):
        if col < 1:  # Các Entry cho cột đầu tiên đến cột thứ ba
            label = tk.Label(row_data_frame, text=labels[col])
            label.grid(row=0, column=col, padx=5, pady=5)
            
            entry = tk.Entry(row_data_frame, width=30)
            entry.grid(row=1, column=col, padx=5, pady=5)
            entry.insert(0, "")  # Khởi tạo ô nhập trống
            entries.append(entry)
            
        elif col < 3:  # Text box cho cột thứ tư
            label = tk.Label(row_data_frame, text=labels[col])
            label.grid(row=0, column=col, padx=5, pady=5)
            
            entry = tk.Entry(row_data_frame, width=30)
            entry.grid(row=1, column=col, padx=5, pady=5)
            entry.insert(0, "")  # Khởi tạo ô nhập trống
            entries.append(entry)
        else:  
            label = tk.Label(row_data_frame, text=labels[col])
            label.grid(row=2, column=col-3, padx=5, pady=5)
            text_box_cavas = tk.Text(row_data_frame, width=75, height=25)
            text_box_cavas.grid(row=3, column=col-3,columnspan=2, padx=10, pady=10)
            text_box = tk.Text(text_box_cavas, width=65, height=25, wrap="word")
            text_box.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            text_box.insert("1.0", "")
            entries.append(text_box)
            # Tạo Scrollbar
            scrollbar = tk.Scrollbar(text_box_cavas, command=text_box.yview)  # Điều khiển theo chiều dọc
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            # Liên kết Text với Scrollbar
            text_box.config(yscrollcommand=scrollbar.set)


    # Nút "сохранить новое упражнение" để thêm hàng mới vào file Excel
    add_data_btn = tk.Button(edit_button_frame, text="сохранить новое упражнение", command=add_row_data)
    add_data_btn.grid(row=0, column=2, padx=5, pady=5)    
        
def edit_row_data():
    global entries  # Đảm bảo danh sách entries được sử dụng đúng cách

    try:
        if not selected_row:
            messagebox.showerror("ошибка", "Ни одна строка не выбрана для редактирования!")
            return

        workbook = load_workbook(selected_file)
        sheet = workbook.active

        # Cập nhật dữ liệu cho hàng đã chọn
        for col, entry in enumerate(entries):
            if isinstance(entry, tk.Entry):  # Nếu là Entry
                value = entry.get()
            elif isinstance(entry, tk.Text):  # Nếu là Text
                value = entry.get("1.0", "end-1c").strip()
            else:
                continue  # Bỏ qua nếu không phải Entry hay Text

            # Ghi dữ liệu vào ô tương ứng, bỏ qua các giá trị trống nếu cần
            sheet.cell(row=selected_row, column=col + 1, value=value or None)

        workbook.save(selected_file)
        workbook.close()  # Đóng workbook để giải phóng file
        select_file(selected_file)
        show_row_data(selected_row)
        messagebox.showinfo("успех", "упражнения успешно обновлены!")
    except Exception as e:
        messagebox.showerror("ошибка", f"Невозможно обновить данные строки:: {e}")


def add_row_data():
    try:
        # Mở workbook đã chọn
        workbook = load_workbook(selected_file)
        sheet = workbook.active

        # Xác định dòng tiếp theo trống trong sheet
        next_row = sheet.max_row + 1

        # Thêm dữ liệu vào từng ô của dòng mới
        for col, entry in enumerate(entries):
            if isinstance(entry, tk.Entry):  # Nếu là Entry (các cột đầu)
                value = entry.get().strip()
            elif isinstance(entry, tk.Text):  # Nếu là Text (các cột có nội dung dài)
                value = entry.get("1.0", "end-1c").strip()
            else:
                continue  # Bỏ qua nếu không phải Entry hay Text

            # Kiểm tra và ghi dữ liệu vào ô, bỏ qua các ô trống nếu cần
            if value:
                sheet.cell(row=next_row, column=col + 1, value=value)

        # Lưu thay đổi vào file Excel
        workbook.save(selected_file)
        workbook.close()  # Đóng workbook sau khi lưu

        # Thông báo thành công
        messagebox.showinfo("успех", "упражнение добавлено успешно!")
        
        # Cập nhật lại giao diện sau khi thêm
        select_file(selected_file)

    except Exception as e:
        messagebox.showerror("ошибка", f"Не удалось добавить упражнение: {e}")

    
    
def clear_row():
    try:
        if selected_row is None:
            messagebox.showwarning("Уведомление", "Пожалуйста, выберите строку для удаления.")
            return

        # Hỏi người dùng подтверждение trước khi xóa
        confirm = messagebox.askyesno("подтверждение", f"Вы уверены, что хотите удалить данные строки? {selected_row}?")
        if not confirm:
            return  # Nếu người dùng nhấn "No", thoát hàm

        workbook = load_workbook(selected_file)
        sheet = workbook.active

        # Xóa dữ liệu của hàng được chọn trong file Excel
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row=selected_row, column=col, value=None)

        workbook.save(selected_file)

        # Xóa widget hiển thị trên giao diện
        for widget in row_data_frame.winfo_children():
            widget.destroy()

        # Hiển thị Уведомление thành công
        messagebox.showinfo("Thành công", f"Dữ liệu của hàng {selected_row} đã được xóa và lưu lại là trống.")

        # Cập nhật lại giao diện
        select_file(selected_file)  # Tải lại file để cập nhật giao diện

    except Exception as e:
        messagebox.showerror("Ошибка", f"Không thể Удаление упражнения: {e}")

def create_new_excel_file():
    new_file_name = new_file_entry.get().strip() + ".xlsx"  # Lấy tên file từ entry
    if not new_file_name:
        messagebox.showerror("ошибка", "Пожалуйста, введите имя оружия.")
        return 

    new_file_path = os.path.join(excel_folder, new_file_name)

# Tạo file Excel mới
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"

    # Thêm tiêu đề cột
    headers = ["имя упражнения", "времия создания", "количество пуля каждого человека", "содержание управнение", "условие упражнения"]
    sheet.append(headers)

    # Lưu file mới
    workbook.save(new_file_path)
    messagebox.showinfo("успех", f"новое оружие: {new_file_name}")

    # Cập nhật danh sách nút
    create_excel_buttons()

# Giao diện Tkinter
root = tk.Tk()
root.geometry('1200x730')
root.title("буййй")
root.grid_rowconfigure(0, minsize=100)  # Đặt chiều cao tối thiểu cho hàng 0
root.grid_columnconfigure(0, minsize=100)

selected_file = None  # Lưu trữ file Excel đã chọn
selected_row = None  # Lưu trữ hàng được chọn
entries = []  # Danh sách Entry và Text box để sửa dữ liệu hàng

def bai_chinh():
    global file_button_frame,row_button_frame,row_data_frame,new_file_entry,edit_button_frame,canvas1,canvas2,fixed_canvas 
    canvas1=tk.Canvas(root,width=125, height=510,bg=bg_color)
    canvas1.place(x=10,y=50)
    
    fixed_canvas_1 = tk.Canvas(root, width=120, height=40, bg=bg_color)
    fixed_canvas_1.place(x=10,y=10)
    fixed_canvas_1.create_text(10,10,text='оружие',anchor="nw",font=("Arial",14, "bold"),fill='white')

    file_button_frame = tk.Frame(canvas1,bg=bg_color)
    canvas1.create_window((0,20),window=file_button_frame,anchor="nw")

    scrollbar1=tk.Scrollbar(root,orient=tk.VERTICAL,command=canvas1.yview)
    scrollbar1.place(x=124,y=10,height=550)
    canvas1.config(yscrollcommand=scrollbar1.set)

    # Frame hiển thị các nút cho từng hàng có dữ liệu
    canvas2=tk.Canvas(root,width=125, height=510,bg=bg_color)
    canvas2.place(x=137,y=50)

    row_button_frame = tk.Frame(canvas2,bg=bg_color )
    canvas2.create_window((0,1),window=row_button_frame,anchor="nw")

    scrollbar2=tk.Scrollbar(root,orient=tk.VERTICAL,command=canvas2.yview)
    scrollbar2.place(x=259,y=10,height=550)
    canvas2.config(yscrollcommand=scrollbar2.set)

    # Frame chứa các nút "исправить данны" và "добовить упражнения"
    edit_button_frame = tk.Frame(root, highlightbackground=bg_color, highlightthickness=3)
    edit_button_frame.place(x=10, y=570, width=600, height=50)  # Dưới row_data_frame

    # Frame chứa ô nhập tên file mới
    new_file_frame = tk.Frame(root, highlightbackground=bg_color, highlightthickness=3)
    new_file_frame.place(x=10, y=640, width=600, height=50) 

    # Entry nhập tên file mới
    new_file_label = tk.Label(new_file_frame, text="новое оружие:")
    new_file_label.grid(row=0, column=0, padx=5, pady=5)
    new_file_entry = tk.Entry(new_file_frame, width=20)
    new_file_entry.grid(row=0, column=1, padx=5, pady=5)

    # Nút để tạo file mới
    create_file_button = tk.Button(new_file_frame, text="новое оружие", command=create_new_excel_file)
    create_file_button.grid(row=0, column=2, padx=5, pady=5)

    # Frame hiển thị dữ liệu của hàng đã chọn hoặc hàng mới
    row_data_frame = tk.Frame(root, highlightbackground=bg_color, highlightthickness=3)
    row_data_frame.place(x=280, y=10, width=900, height=540)  # Chuyển sang bên phải các nút


    # Tạo Canvas cố định (cho text_id)
    fixed_canvas = tk.Canvas(root, width=120, height=40, bg=bg_color)
    fixed_canvas.place(x=137,y=10)


    # Tạo các nút từ các file Excel
    create_excel_buttons()


import sqlite3
conn = sqlite3.connect('users.db')
cursor = conn.cursor()
locked_icon= tk.PhotoImage(file='./images/locked.png')
unlocked_icon= tk.PhotoImage(file='./images/unlocked.png')
icon_1= tk.PhotoImage(file='./images/login1.png')
icon_2= tk.PhotoImage(file='./images/login_user.png')
icon_3= tk.PhotoImage(file='./images/login2.png')
add_student_icon = tk.PhotoImage(file='./images/login3.png')

# Kiểm tra nếu bảng 'users' chưa tồn tại, thì tạo mới
cursor.execute('''
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY,
        username TEXT,
        password TEXT
    )
''')
conn.commit()


def confirmation_box(message):

    answer= tk.BooleanVar()
    answer.set(False)


    def action(ans):
        answer.set(ans)
        confirmation_box_fm.destroy()



    confirmation_box_fm=tk.Frame(root, highlightbackground=bg_color,
                                 highlightthickness=3)
    
    message_lb=tk.Label(confirmation_box_fm,text=message, font=('bold',15))
    message_lb.pack(pady=20)

    cancel_lb=tk.Button(confirmation_box_fm,text='Cancel', font=('Bold',15),
                        bd=0, bg=bg_color,fg='white',
                        command=lambda: action(False))
    cancel_lb.place(x=50, y=160, width=80)

    yes_lb=tk.Button(confirmation_box_fm,text='Yes', font=('Bold',15),
                        bd=0, bg=bg_color,fg='white',
                        command=lambda: action(True))
    yes_lb.place(x=190, y=160, width=80)

    confirmation_box_fm.place(x=100, y=120 , width=320, height=220)

    root.wait_window(confirmation_box_fm)
    return answer.get()
    
    

def message_box():
    message_box_fm=tk.Frame(root, highlightbackground=bg_color,
                                 highlightthickness=3)
    
    close_btn= tk.Button(message_box_fm, text='x',bd=0,font=())

    close_btn.place(x=100, y=120, width=320, height=200)

def welcome_page():

    def forward_to_student_login_page():
        welcome_page_fm.destroy()
        root.update()
        student_login_page()

    def forward_to_admin_login_page():
        welcome_page_fm.destroy()
        root.update()
        admin_login_page()
    

    welcome_page_fm = tk.Frame(root, highlightbackground=bg_color, highlightthickness=3)

    heading_lb = tk.Label(welcome_page_fm,
                        text='Добро пожаловать в программное \n обеспечение для управления оружием' ,
                        bg=bg_color, fg='white', font=('bold', 18), anchor='w')
    heading_lb.place(x=0, y=0, width=800)



    student_login_btn = tk.Button(welcome_page_fm, text='Вход', bg=bg_color,
                                fg='white', font=('bold',15), bd=0,
                                command=forward_to_student_login_page)
    student_login_btn.place(x=130, y=125 ,width=200)

    student_login_img = tk.Button(welcome_page_fm,image=icon_1, bd=0,
                                  command=forward_to_student_login_page)
    student_login_img.place(x=70, y=100 )


    admin_login_btn = tk.Button(welcome_page_fm, text='Регистрация', bg=bg_color,
                                fg='white', font=('bold',15), bd=0,
                                command=forward_to_admin_login_page)
    admin_login_btn.place(x=130, y=225 ,width=200)

    admin_login_img = tk.Button(welcome_page_fm,image=icon_2, bd=0,
                                command=forward_to_admin_login_page)
    admin_login_img.place(x=70, y=200 )


    welcome_page_fm.pack(pady=30)
    welcome_page_fm.pack_propagate(False)
    welcome_page_fm.configure(width=460, height=420)



def student_login_page():

    # Hàm xử lý đăng nhập
    def login_user():
        username = id_number_ent.get()
        password = password__ent.get()
        
        cursor.execute("SELECT * FROM users WHERE username=? AND password=?", (username, password))
        user = cursor.fetchone()
        if user:
            messagebox.showinfo("Войти успешно", "Добро пожаловать на главный экран!")
            student_login_page_fm.destroy()
            root.update()
            bai_chinh()


            # Đăng nhập thành công, có thể thực hiện các thao tác tiếp theo ở đây
        else:
            messagebox.showerror("Ошибка входа", "Неверное имя пользователя или пароль")
    def show_hide_password():
        if password__ent['show']=='*':
            password__ent.config(show='')
            show_hide_btn.config(image=unlocked_icon)
        else:
            password__ent.config(show='*')
            show_hide_btn.config(image=locked_icon)

    def forward_to_welcome_page():
        student_login_page_fm.destroy()
        root.update()
        welcome_page()
    
    student_login_page_fm= tk.Frame(root, highlightbackground=bg_color, highlightthickness=3)


    heading_lb= tk.Label(student_login_page_fm, text='Войти Студент' ,
                            bg=bg_color, fg='white', font=('bold', 18))

    heading_lb.place(x=0, y=0, width=400)


    back_btn= tk.Button(student_login_page_fm, text='←', font=('bold',20),
                        fg=bg_color,bd=0,
                        command=forward_to_welcome_page)
    back_btn.place(x=5,y=40)

    stud_icon_lb= tk.Label(student_login_page_fm, image=icon_3)
    stud_icon_lb.place(x=150, y=40)

    id_number_lb=tk.Label(student_login_page_fm, text='введите имя курсанта.',
                        font=('bold', 15), fg=bg_color)
    id_number_lb.place(x=80,y=140)


    id_number_ent= tk.Entry(student_login_page_fm,font=('bold', 15),
                            justify=tk.CENTER, 
                            highlightcolor=bg_color, 
                            highlightbackground='gray',
                            highlightthickness=2)
    id_number_ent.place(x=80,y=190)


    password_lb=tk.Label(student_login_page_fm, text='введите пароль курсанта.',
                        font=('bold', 15), fg=bg_color)
    password_lb.place(x=80,y=240)

    password__ent= tk.Entry(student_login_page_fm,font=('bold', 15),
                            justify=tk.CENTER, 
                            highlightcolor=bg_color, 
                            highlightbackground='gray',
                            highlightthickness=2,
                            show='*')
    password__ent.place(x=80,y=290)


    show_hide_btn= tk.Button(student_login_page_fm, image=locked_icon, bd=0,
                            command=show_hide_password)
    show_hide_btn.place(x=310,y=280)

    login_bth= tk.Button(student_login_page_fm, text='логин',
                        font=('Bold',15), bg=bg_color,fg='white',command=login_user)
    login_bth.place(x=95, y=340, width=200, height=40)


    forget_password_btn= tk.Button(student_login_page_fm, text='\n Забыли пароль?',
                                fg=bg_color, bd=0)
    forget_password_btn.place(x=150, y=390)



    student_login_page_fm.pack(pady=30)
    student_login_page_fm.pack_propagate(False)
    student_login_page_fm.configure(width=400, height=450)


def admin_login_page():

    def register_user():
        new_username = id_number_ent.get()
        new_password = password__ent.get()
    
        cursor.execute("SELECT * FROM users WHERE username=?", (new_username,))
        existing_user = cursor.fetchone()
        if existing_user:
            messagebox.showerror("Ошибка", "Имя пользователя уже существует")
        else:
            cursor.execute("INSERT INTO users (username, password) VALUES (?, ?)", (new_username, new_password))
            conn.commit()
            messagebox.showinfo("Успех", "Регистрация прошла успешно")

    def show_hide_password():
        if password__ent['show']=='*':
            password__ent.config(show='')
            show_hide_btn.config(image=unlocked_icon)
        else:
            password__ent.config(show='*')
            show_hide_btn.config(image=locked_icon)

    def forward_to_welcome_page():
        admin_login_page_fm.destroy()
        root.update()
        welcome_page()

    admin_login_page_fm= tk.Frame(root, highlightbackground=bg_color, highlightthickness=3)


    heading_lb= tk.Label(admin_login_page_fm, text='Регистрация' ,
                            bg=bg_color, fg='white', font=('bold', 18))

    heading_lb.place(x=0, y=0, width=400)


    back_btn= tk.Button(admin_login_page_fm, text='←', font=('bold',20),
                        fg=bg_color,bd=0,
                        command=forward_to_welcome_page)
    back_btn.place(x=5,y=40)

    admin_icon_lb= tk.Label(admin_login_page_fm, image=add_student_icon)
    admin_icon_lb.place(x=150, y=40)

    id_number_lb=tk.Label(admin_login_page_fm, text='введите имя пользователя.',
                        font=('bold', 15), fg=bg_color)
    id_number_lb.place(x=80,y=140)


    id_number_ent= tk.Entry(admin_login_page_fm,font=('bold', 15),
                            justify=tk.CENTER, 
                            highlightcolor=bg_color, 
                            highlightbackground='gray',
                            highlightthickness=2)
    id_number_ent.place(x=80,y=190)


    password_lb=tk.Label(admin_login_page_fm, text='введите пароль пользователя.',
                        font=('bold', 15), fg=bg_color)
    password_lb.place(x=80,y=240)

    password__ent= tk.Entry(admin_login_page_fm,font=('bold', 15),
                            justify=tk.CENTER, 
                            highlightcolor=bg_color, 
                            highlightbackground='gray',
                            highlightthickness=2,
                            show='*')
    password__ent.place(x=80,y=290)


    show_hide_btn= tk.Button(admin_login_page_fm, image=locked_icon, bd=0,
                            command=show_hide_password)
    show_hide_btn.place(x=310,y=280)

    login_bth= tk.Button(admin_login_page_fm, text='Зарегистрироваться',
                        font=('Bold',15), bg=bg_color,fg='white', command=register_user)
    login_bth.place(x=95, y=340, width=200, height=40)




    admin_login_page_fm.pack(pady=30)
    admin_login_page_fm.pack_propagate(False)
    admin_login_page_fm.configure(width=400, height=450)

welcome_page()
root.mainloop()